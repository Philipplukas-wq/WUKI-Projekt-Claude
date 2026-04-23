"""
WU-Berater: Export-Skript für unterjährige Wirtschaftlichkeitsuntersuchungen
Füllt das Excel-Template 'Template Dokumentation WU unterjährig.xlsm' mit WU-Inhalten.

Verwendung:
    from export_wu_unterjahrig import fill_template
    fill_template(wu_data, output_path)
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import date
import os
from wu_file_handler import cleanup_lock_files

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'Template Dokumentation WU unterjährig.xlsm')
TEMPLATE_PATH = os.path.normpath(TEMPLATE_PATH)

OUTPUT_DIR = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', '..', '..', 'Erstellte WU', 'Unterjährig'))


def fill_template(wu_data: dict, output_path: str):
    """
    Befüllt das unterjährige WU-Template mit den übergebenen Daten.

    :param wu_data: Dictionary mit WU-Inhalten (Struktur siehe WU_DATA_SCHEMA)
    :param output_path: Pfad der Ausgabedatei (.xlsm)
    """
    wb = openpyxl.load_workbook(TEMPLATE_PATH, keep_vba=True)
    ws = wb['WU Vermerk']
    meta = wu_data.get('meta', {})
    inhalt = wu_data.get('inhalt', {})

    # Kopfdaten
    ws['B6'] = meta.get('dienststelle', '')
    ws['F6'] = meta.get('bearbeiter', '')
    ws['B7'] = _parse_date(meta.get('datum', ''))
    ws['D7'] = _parse_date(meta.get('beginn_massnahme', meta.get('datum', '')))
    ws['F4'] = f"{meta.get('schutz', 'offen')}\nVersion: {meta.get('datum', '')}"

    # Vermögenstyp (A5/D5 mit großer Checkbox)
    vermögenstyp = inhalt.get('vermögenstyp', '')
    if vermögenstyp == 'Anlagevermögen':
        ws['A5'] = 'X'
        ws['A5'].font = Font(size=24, bold=True)
    elif vermögenstyp == 'Umlaufvermögen':
        ws['D5'] = 'X'
        ws['D5'].font = Font(size=24, bold=True)

    # Bedarfsforderung
    ws['B8'] = inhalt.get('bedarfsforderung', '')

    # Zeilenhöhen anpassen
    ws.row_dimensions[8].height = 80
    ws.row_dimensions[15].height = 60
    ws.row_dimensions[19].height = 60

    # Häkchen setzen
    checkmark_font = Font(size=11, bold=True, color='1F3864')
    center_align = Alignment(horizontal='center', vertical='center')

    haken = inhalt.get('haken', {})
    haken_map = {
        'kauf_benoetigt':              'A10',
        'bedarf_neu':                  'A11',
        'sonstiges_bedarf':            'A12',
        'eigenleistung_grund1':        'A14',
        'eigenleistung_grund2':        'A15',
        'miete_grund1':                'A17',
        'miete_grund2':                'A18',
        'miete_grund3':                'A19',
        'keine_folgeausgaben':         'A21',
    }
    for key, cell_addr in haken_map.items():
        if haken.get(key, False):
            ws[cell_addr] = 'X'
            ws[cell_addr].font = checkmark_font
            ws[cell_addr].alignment = center_align

    # Freitextfelder
    if inhalt.get('eigenleistung_begruendung'):
        ws['F15'] = inhalt['eigenleistung_begruendung']
    if inhalt.get('miete_begruendung'):
        ws['F19'] = inhalt['miete_begruendung']

    # Voraussichtliche Ausgaben
    if inhalt.get('ausgaben'):
        ws['F22'] = inhalt['ausgaben']

    # Anlage-Blatt mit Hyperlinks
    anlage = wu_data.get('anlage', [])
    if anlage:
        _create_anlage_sheet(wb, anlage, meta.get('datum', ''))

    # Räume Lock-Dateien auf (Excel-Dateien)
    output_dir = os.path.dirname(output_path)
    cleanup_lock_files(output_dir)

    wb.save(output_path)
    print(f'[OK] WU-Dokument gespeichert: {os.path.basename(output_path)}')
    return output_path


def _parse_date(datum_str: str):
    """Konvertiert 'TT.MM.JJJJ' in date-Objekt."""
    try:
        teile = datum_str.split('.')
        return date(int(teile[2]), int(teile[1]), int(teile[0]))
    except Exception:
        return datum_str


def _create_anlage_sheet(wb, anlage: list, datum: str):
    from openpyxl.styles import Font as XFont, Alignment as XAlign

    if 'Anlagen' in wb.sheetnames:
        del wb['Anlagen']

    ws = wb.create_sheet('Anlagen')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = XFont(bold=True, color='FFFFFF', size=12)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    ws['A1'] = 'ANLAGEN: Quellenangaben'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:D1')

    ws['A3'] = 'Quelle'
    ws['B3'] = 'Datum'
    ws['C3'] = 'Ergebnis'
    ws['D3'] = 'Bemerkung'

    for cell in ['A3', 'B3', 'C3', 'D3']:
        ws[cell].font = XFont(bold=True)
        ws[cell].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws[cell].border = thin_border

    row = 4
    for quelle in anlage:
        quelle_text = quelle.get('quelle', quelle.get('produkt', ''))
        quelle_link = quelle.get('link', quelle.get('url', ''))

        # Quelle mit Hyperlink
        if quelle_link:
            ws[f'A{row}'] = quelle_text
            ws[f'A{row}'].hyperlink = quelle_link
            ws[f'A{row}'].font = XFont(underline='single', color='0563C1')
        else:
            ws[f'A{row}'] = quelle_text

        ws[f'B{row}'] = quelle.get('datum', '')
        ws[f'C{row}'] = quelle.get('ergebnis', quelle.get('preis', ''))
        ws[f'D{row}'] = quelle.get('bemerkung', '')

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = thin_border
        row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 35


def erstelle_abschlusscheckliste_unterjahrig(wu_data: dict, output_path: str) -> str:
    """
    Gibt die Abschlusscheckliste für eine unterjährige WU zurück.
    """
    meta   = wu_data.get('meta', {})
    inhalt = wu_data.get('inhalt', {})
    anlage = wu_data.get('anlage', [])

    haken = inhalt.get('haken', {})
    gesetzte_haken = [k for k, v in haken.items() if v]

    auto_befuellt = [
        f'Dokumentkopf (Dienststelle: {meta.get("dienststelle", "")}, '
        f'Bearbeiter: {meta.get("bearbeiter", "")}, Schutz: offen)',
        'Bedarfsforderung (qualitativ + quantitativ)',
        f'Bisherige Bedarfsdeckung (Haken: {", ".join(gesetzte_haken)})',
        'Ausschluss Eigenleistung (Begründung)',
        'Ausschluss Miete/Leasing (inkl. Preisvergleich)',
        'Bestätigung Unterjährigkeit (keine Folgeausgaben)',
        f'Voraussichtliche Ausgaben: {inhalt.get("ausgaben", "–")} €',
        f'Anlage Marktrecherche: {len(anlage)} Quellen (separates Blatt)',
    ]

    manuell_erforderlich = [
        'Checkboxen in der Excel-Datei manuell setzen (Klick auf die Kontrollkästchen)',
        'Screenshots der Marktrecherche-Webseiten in Anlage einfügen',
        'Bearbeiter/-in prüfen (Zelle F6)',
    ]

    checkliste = (
        f'\n{"="*55}\n'
        f'ABSCHLUSS-CHECKLISTE (unterjährige WU)\n'
        f'Datei: {output_path}\n'
        f'{"="*55}\n\n'
        f'Automatisch befüllt:\n'
    )
    for item in auto_befuellt:
        checkliste += f'  [x] {item}\n'
    checkliste += '\nNoch manuell zu ergänzen:\n'
    for item in manuell_erforderlich:
        checkliste += f'  [ ] {item}\n'
    checkliste += f'\n{"="*55}\n'
    return checkliste


def build_filename(datum: str, sachverhalt: str, dienststelle: str, version: int = 1) -> str:
    """Vollständiger Ausgabepfad: <OUTPUT_DIR>/JJJJMMTT_WU_[Sachverhalt]_[Dienststelle]_Version_N.xlsm"""
    teile = datum.split('.')
    jjmmtt = teile[2] + teile[1] + teile[0] if len(teile) == 3 else datum.replace('.', '')
    s = sachverhalt.replace(' ', '_').replace('/', '_')
    d = dienststelle.replace(' ', '_').replace('/', '_')
    filename = f"{jjmmtt}_WU_{s}_{d}_Version_{version}.xlsm"
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    return os.path.join(OUTPUT_DIR, filename)


WU_DATA_SCHEMA = {
    "meta": {
        "dienststelle":      "BAIUDBw ...",
        "bearbeiter":        "Vorname Nachname",
        "datum":             "TT.MM.JJJJ",
        "beginn_massnahme":  "TT.MM.JJJJ",
        "schutz":            "offen",
        "version":           "1",
    },
    "inhalt": {
        "vermögenstyp":                "Anlagevermögen|Umlaufvermögen",
        "bedarfsforderung":            "Funktionale, lösungsneutrale Bedarfsbeschreibung ...",
        "haken": {
            "kauf_benoetigt":          True,
            "bedarf_neu":              False,
            "sonstiges_bedarf":        False,
            "eigenleistung_grund1":    False,    # A14 — Grund 1
            "eigenleistung_grund2":    True,     # A15 — Sonstiges/Grund 2 (BUNDESWEHR: immer mindestens eine Grund-Checkbox)
            "miete_grund1":            False,    # A17 — Grund 1
            "miete_grund2":            False,    # A18 — Grund 2
            "miete_grund3":            True,     # A19 — Sonstiges/Grund 3 (wenn Miete ausgeschlossen)
            "keine_folgeausgaben":     True,
        },
        "eigenleistung_begruendung":   "kein geeignetes Gerät im Eigenbestand ...",
        "miete_begruendung":           "wirtschaftlich nicht vorteilhaft ...",
        "ausgaben":                    205.00,
    },
    "anlage": [
        {"quelle": "Produkt XYZ", "datum": "TT.MM.JJJJ", "ergebnis": "Preis", "bemerkung": "...", "link": "https://..."},
    ],
}
