from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime
import os

template_path = r"P:\WUKI_Projekt\Claude\Template Dokumentation WU unterjährig Vermerk.xlsm"
output_dir = r"P:\WUKI_Projekt\Claude\Erstellte WU\Unterjährig"

def safe_set_cell_value(ws, cell_ref, value):
    try:
        ws[cell_ref].value = value
    except AttributeError:
        pass

def create_document(beispiel_name, daten):
    wb = load_workbook(template_path, keep_vba=True)
    ws = wb.active

    # ZEILENHÖHEN ANPASSEN
    ws.row_dimensions[15].height = 60  # Eigenleistungs-Begründung
    ws.row_dimensions[19].height = 60  # Miete/Leasing-Begründung

    # METADATEN
    safe_set_cell_value(ws, 'B6', daten['dienststelle'])
    safe_set_cell_value(ws, 'F6', daten['bearbeiter'])
    safe_set_cell_value(ws, 'B7', daten['datum'])
    safe_set_cell_value(ws, 'D7', daten['massnahmenbeginn'])

    # VERMÖGENSTYP - Kreuz größer formatieren
    if daten['vermögenstyp'] == 'Anlagevermögen':
        safe_set_cell_value(ws, 'A5', 'X')
        ws['A5'].font = Font(size=24, bold=True)
    else:
        safe_set_cell_value(ws, 'D5', 'X')
        ws['D5'].font = Font(size=24, bold=True)

    # BEDARFSFORDERUNG
    safe_set_cell_value(ws, 'B8', daten['bedarfsforderung'])

    # CHECKBOXEN BEDARF
    bedarf_option = daten.get('bedarf_option', 1)
    if bedarf_option == 1:
        safe_set_cell_value(ws, 'A10', 'X')
        safe_set_cell_value(ws, 'A11', ' ')
        safe_set_cell_value(ws, 'A12', ' ')
    elif bedarf_option == 2:
        safe_set_cell_value(ws, 'A10', ' ')
        safe_set_cell_value(ws, 'A11', 'X')
        safe_set_cell_value(ws, 'A12', ' ')
    elif bedarf_option == 3:
        safe_set_cell_value(ws, 'A10', ' ')
        safe_set_cell_value(ws, 'A11', ' ')
        safe_set_cell_value(ws, 'A12', 'X')
        if daten.get('bedarf_begruendung'):
            safe_set_cell_value(ws, 'F12', daten['bedarf_begruendung'])

    # AUSSCHLUSS EIGENLEISTUNG
    # A13: TEXT BLEIBT STEHEN - KEIN KREUZ!
    # Kreuze nur in A14 oder A15 (für die Gründe)
    safe_set_cell_value(ws, 'A14', daten.get('checkbox_eigenleistung_grund1', ' '))
    safe_set_cell_value(ws, 'A15', daten.get('checkbox_eigenleistung_grund2', ' '))
    # Wenn A15 angehakt: Begründung in F15
    if daten.get('checkbox_eigenleistung_grund2') == 'X':
        if daten.get('eigenleistung_begruendung'):
            safe_set_cell_value(ws, 'F15', daten['eigenleistung_begruendung'])
    # AUSSCHLUSS MIETE/LEASING
    # A16: TEXT BLEIBT STEHEN - KEIN KREUZ!
    # Kreuze nur in A17, A18, A19 (für die Gründe)
    safe_set_cell_value(ws, 'A17', daten.get('checkbox_miete_grund1', ' '))
    safe_set_cell_value(ws, 'A18', daten.get('checkbox_miete_grund2', ' '))
    safe_set_cell_value(ws, 'A19', daten.get('checkbox_miete_grund3', ' '))
    # Wenn A19 angehakt: Begründung in F19
    if daten.get('checkbox_miete_grund3') == 'X':
        if daten.get('miete_begruendung'):
            safe_set_cell_value(ws, 'F19', daten['miete_begruendung'])

    # sonstiges_text wird nicht mehr benötigt - Begründung geht in F19

    # UNTERJAEHRIGKEIT
    safe_set_cell_value(ws, 'A21', daten.get('checkbox_unterjaehrigkeit', ' '))

    # F22: KALKULIERTER PREIS (WICHTIG: F22, NICHT G22!)
    safe_set_cell_value(ws, 'F22', daten.get('kaufpreis', 0))

    # ANLAGEN BLATT
    if 'Anlagen' not in wb.sheetnames:
        anlagen_ws = wb.create_sheet('Anlagen')
    else:
        anlagen_ws = wb['Anlagen']

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    anlagen_ws['A1'] = 'ANLAGEN: Quellenangaben'
    anlagen_ws['A1'].font = header_font
    anlagen_ws['A1'].fill = header_fill
    anlagen_ws.merge_cells('A1:D1')

    anlagen_ws['A3'] = 'Quelle'
    anlagen_ws['B3'] = 'Datum'
    anlagen_ws['C3'] = 'Ergebnis'
    anlagen_ws['D3'] = 'Bemerkung'

    for cell in ['A3', 'B3', 'C3', 'D3']:
        anlagen_ws[cell].font = Font(bold=True)
        anlagen_ws[cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        anlagen_ws[cell].border = thin_border

    row = 4
    for quelle in daten.get('quellen', []):
        # Quelle mit Hyperlink
        quelle_text = quelle.get('quelle', '')
        quelle_link = quelle.get('link', '')
        if quelle_link:
            anlagen_ws[f'A{row}'] = quelle_text
            anlagen_ws[f'A{row}'].hyperlink = quelle_link
            anlagen_ws[f'A{row}'].font = Font(underline='single', color='0563C1')
        else:
            anlagen_ws[f'A{row}'] = quelle_text

        anlagen_ws[f'B{row}'] = quelle.get('datum', '')
        anlagen_ws[f'C{row}'] = quelle.get('ergebnis', '')
        anlagen_ws[f'D{row}'] = quelle.get('bemerkung', '')
        for col in ['A', 'B', 'C', 'D']:
            anlagen_ws[f'{col}{row}'].border = thin_border
        row += 1

    anlagen_ws.column_dimensions['A'].width = 40
    anlagen_ws.column_dimensions['B'].width = 15
    anlagen_ws.column_dimensions['C'].width = 25
    anlagen_ws.column_dimensions['D'].width = 35

    return wb

# BEISPIEL 1
daten1 = {
    'dienststelle': 'KompZWuBw',
    'bearbeiter': 'Philipp Lukas',
    'datum': datetime.now(),
    'massnahmenbeginn': datetime.strptime('23.04.2026', '%d.%m.%Y'),
    'vermögenstyp': 'Umlaufvermögen',
    'bedarfsforderung': 'Befestigungsmittel zur Verschraubung von Stahlkonstruktionen. Gewindedurchmesser M8, Länge 20mm, Kopfform Senkkopf, Material Stahl verzinkt. Menge: 500 Stück.',
    'bedarf_option': 1,
    'checkbox_eigenleistung_grund1': 'X',
    'checkbox_eigenleistung_grund2': ' ',
    'checkbox_miete_grund1': 'X',
    'checkbox_miete_grund2': ' ',
    'checkbox_miete_grund3': ' ',
    'checkbox_unterjaehrigkeit': 'X',
    'kaufpreis': 20.00,
    'quellen': [
        {'quelle': 'Normteile Leinigen: Senkschraube M8x20', 'datum': '23.04.2026',
         'ergebnis': '20,00 EUR (500 Stück)', 'bemerkung': 'Marktüblicher Preis',
         'link': 'https://www.normteile-leinigen.de/Senkschraube-M8x20-Stahl-verzinkt-8.8-DIN-EN-ISO-10642/10642108020088'}
    ]
}

wb1 = create_document('Schrauben', daten1)
filename1 = os.path.join(output_dir, "20260423_WU_Schrauben_KompZWuBw_AUV_Version_1.xlsm")
wb1.save(filename1)
print("OK: Schrauben")

# BEISPIEL 2
daten2 = {
    'dienststelle': 'KompzWuBw',
    'bearbeiter': 'Philipp Lukas',
    'datum': datetime.now(),
    'massnahmenbeginn': datetime.strptime('23.04.2026', '%d.%m.%Y'),
    'vermögenstyp': 'Anlagevermögen',
    'bedarfsforderung': 'Akkubetriebenes Elektrowerkzeug zum Bohren mit Schnellspannfutter. Mindestanforderungen: Drehmoment bis 60 Nm. Einsatz für tägliche Wartungsarbeiten (ca. 220 Tage/Jahr, 3-4 Stunden pro Einsatztag). Eigenständige Verfügbarkeit erforderlich.',
    'bedarf_option': 2,
    # Eigenleistungs-Ausschluss: A15 (Sonstiges) mit Begründung
    'checkbox_eigenleistung_grund1': ' ',  # A14 leer
    'checkbox_eigenleistung_grund2': 'X',  # A15 angehakt (Sonstiges)
    'eigenleistung_begruendung': 'Werkzeug muss eigenständig verfügbar sein',  # F15
    'checkbox_miete_grund3': 'X',
    'miete_begruendung': 'Wartungsarbeiten erfordern sofortige Verfügbarkeit',
    'checkbox_unterjaehrigkeit': 'X',
    'kaufpreis': 189.50,
    'quellen': [
        {'quelle': 'Scheppach: Akku-Bohrschrauber BC-DD60-X', 'datum': '23.04.2026',
         'ergebnis': 'ab 189,50 EUR', 'bemerkung': '60 Nm Drehmoment, Brushless',
         'link': 'https://shop.scheppach.com/Akku-Bohrschrauber-BC-DD60-X-Scheppach-Brushless-Drehmoment-60Nm-Drehmomentstufen-25-1-LED/5909239900'},
        {'quelle': 'Wall Baumaschinen: Akkubohrschrauber Tagesmiete', 'datum': '23.04.2026',
         'ergebnis': '12,50 EUR/Tag (netto)', 'bemerkung': 'Break-even: 15 Tage',
         'link': 'https://wall-baumaschinen.de/Maschinen/akkubohrschrauber-mieten/'}
    ]
}

wb2 = create_document('Bohrmaschine', daten2)
filename2 = os.path.join(output_dir, "20260423_WU_Bohrmaschine_KompzWuBw_AUV_Version_2.xlsm")
wb2.save(filename2)
print("OK: Bohrmaschine")

print("\nBEIDE DOKUMENTE ERSTELLT:")
print("- F22: Kalkulierter Kaufpreis eingetragen")
print("- Anlagen-Blatt mit Quellenangaben erstellt")
